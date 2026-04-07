from __future__ import annotations

from collections.abc import Mapping
from typing import Any, cast
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from receipt_app.config import AppConfig, DEFAULT_CONFIG, GUIDE_NOTICE_TEXT
from receipt_app.models import ExportRow, UploadedReceipt
from receipt_app.utils.images import (
    image_to_png_bytes,
    open_image_from_bytes,
    resize_for_excel,
)

THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
MALGUN_FONT = Font(name="Malgun Gothic", size=12)
CALIBRI_FONT = Font(name="Calibri", size=12)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_VCENTER = Alignment(horizontal="left", vertical="center")
VCENTER = Alignment(vertical="center")

FILL_GREEN = PatternFill(patternType="solid", fgColor=Color(indexed=11))
FILL_WHITE = PatternFill(patternType="solid", fgColor=Color(indexed=9))
FILL_BLUE = PatternFill(patternType="solid", fgColor=Color(indexed=12))


def build_workbook_bytes(
    person_name: str,
    rows: list[ExportRow | Mapping[str, object]],
    receipts: list[UploadedReceipt],
    template_path: str | Path | None = None,
    month: date | str | None = None,
    config: AppConfig = DEFAULT_CONFIG,
) -> bytes:
    normalized_rows = _coerce_rows(rows)
    report_month = month or date.today().replace(day=1)
    return export_receipts_to_workbook(
        rows=normalized_rows,
        receipts=receipts,
        operator_name=person_name,
        report_month_text=report_month,
        template_path=template_path,
        config=config,
    )


def export_receipts_to_workbook(
    rows: list[ExportRow],
    receipts: list[UploadedReceipt],
    operator_name: str,
    report_month_text: date | str,
    template_path: str | Path | None = None,
    config: AppConfig = DEFAULT_CONFIG,
) -> bytes:
    workbook = _create_workbook(config)
    sheet = workbook[config.sheet_name]

    _ = template_path
    sheet[config.operator_name_cell] = operator_name
    sheet[config.total_amount_cell] = "=SUM(E7:E28)"

    report_month = _coerce_report_month(report_month_text)
    sheet[config.month_cell] = report_month
    if isinstance(report_month, date):
        sheet[config.month_cell].number_format = "mmm-yy"

    _write_rows(sheet=sheet, rows=rows, config=config)
    _embed_receipt_images(sheet=sheet, receipts=receipts, config=config)

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _create_workbook(config: AppConfig) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = config.sheet_name

    _configure_main_sheet(sheet, config)
    _create_guide_sheet(workbook, config)
    return workbook


def _configure_main_sheet(sheet: Worksheet, config: AppConfig) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_format.defaultColWidth = 11.1667
    sheet.sheet_format.baseColWidth = 8

    for column, width in {
        "A": 5.0,
        "C": 12.1719,
        "F": 4.85156,
        "G": 38.8516,
        "L": 11.1719,
        "M": 11.1719,
    }.items():
        sheet.column_dimensions[column].width = width

    sheet.row_dimensions[1].height = 18.0
    for row_number in range(2, 30):
        sheet.row_dimensions[row_number].height = 24.0
    for row_number in range(30, 111):
        sheet.row_dimensions[row_number].height = 18.0

    _style_main_background(sheet)
    _style_header_area(sheet, config)
    _style_table_area(sheet, config)
    _style_image_gallery(sheet, config)


def _style_main_background(sheet: Worksheet) -> None:
    for row_number in range(1, 111):
        for column in "ABCDEFGHIJKL":
            cell = sheet[f"{column}{row_number}"]
            cell.font = CALIBRI_FONT
            cell.border = THIN_BORDER
            cell.fill = FILL_WHITE
            cell.alignment = VCENTER

    for column in ("G", "L"):
        cell = sheet[f"{column}1"]
        cell.border = Border(
            left=THIN_SIDE,
            right=THIN_SIDE,
            top=THIN_SIDE,
            bottom=Side(style=None),
        )

    for row_number in range(2, 110):
        l_cell = sheet[f"L{row_number}"]
        l_cell.border = Border(
            left=THIN_SIDE,
            right=THIN_SIDE,
            top=Side(style=None),
            bottom=Side(style=None),
        )
    sheet["L110"].border = Border(
        left=THIN_SIDE,
        right=THIN_SIDE,
        top=Side(style=None),
        bottom=THIN_SIDE,
    )


def _style_header_area(sheet: Worksheet, config: AppConfig) -> None:
    header_labels = {"B2": "이름", "B3": "총액", "B4": "회차"}
    for coord, value in header_labels.items():
        cell = sheet[coord]
        cell.value = value
        cell.font = MALGUN_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.fill = FILL_GREEN
        cell.number_format = "@"

    for coord in (
        config.operator_name_cell,
        config.total_amount_cell,
        config.month_cell,
    ):
        cell = sheet[coord]
        cell.font = MALGUN_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.fill = FILL_WHITE
    sheet[config.operator_name_cell].number_format = "@"

    for row_number in (5,):
        for column in ("B", "C", "D", "E"):
            cell = sheet[f"{column}{row_number}"]
            cell.font = CALIBRI_FONT
            cell.border = THIN_BORDER
            cell.fill = FILL_WHITE
            cell.alignment = VCENTER

    for row_number in (2, 3, 4):
        cell = sheet[f"E{row_number}"]
        cell.font = CALIBRI_FONT
        cell.border = THIN_BORDER
        cell.fill = FILL_WHITE
        cell.alignment = VCENTER

    for merged_range in config.header_value_merge_ranges:
        sheet.merge_cells(merged_range)


def _style_table_area(sheet: Worksheet, config: AppConfig) -> None:
    table_header_labels = {"B6": "번호", "C6": "비목", "D6": "세목", "E6": "금액"}
    for coord, value in table_header_labels.items():
        cell = sheet[coord]
        cell.value = value
        cell.font = MALGUN_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.fill = FILL_GREEN
        cell.number_format = "@"

    for row_number in range(config.table_start_row, config.table_border_end_row + 1):
        b_cell = sheet[f"B{row_number}"]
        b_cell.font = (
            MALGUN_FONT if row_number <= config.numbered_table_end_row else CALIBRI_FONT
        )
        b_cell.border = THIN_BORDER
        b_cell.alignment = VCENTER
        b_cell.fill = FILL_WHITE

        c_cell = sheet[f"C{row_number}"]
        c_cell.font = (
            MALGUN_FONT if row_number <= config.numbered_table_end_row else CALIBRI_FONT
        )
        c_cell.border = THIN_BORDER
        c_cell.alignment = LEFT_VCENTER if row_number <= 11 else VCENTER
        c_cell.fill = FILL_WHITE
        if row_number <= 11:
            c_cell.number_format = "@"

        d_cell = sheet[f"D{row_number}"]
        d_cell.font = (
            MALGUN_FONT if row_number <= config.numbered_table_end_row else CALIBRI_FONT
        )
        d_cell.border = THIN_BORDER
        d_cell.alignment = LEFT_VCENTER if row_number <= 11 else VCENTER
        d_cell.fill = FILL_WHITE
        if row_number <= 11:
            d_cell.number_format = "@"

        e_cell = sheet[f"E{row_number}"]
        e_cell.font = (
            MALGUN_FONT if row_number <= config.numbered_table_end_row else CALIBRI_FONT
        )
        e_cell.border = THIN_BORDER
        e_cell.alignment = VCENTER
        e_cell.fill = FILL_WHITE

    for index, row_number in enumerate(
        range(config.table_start_row, config.numbered_table_end_row + 1), start=1
    ):
        sheet[f"B{row_number}"] = index


def _style_image_gallery(sheet: Worksheet, config: AppConfig) -> None:
    g2_border = Border(
        left=THIN_SIDE,
        right=THIN_SIDE,
        top=Side(style=None),
        bottom=THIN_SIDE,
    )
    g2_cell = sheet["G2"]
    g2_cell.value = "영수증사진"
    g2_cell.font = MALGUN_FONT
    g2_cell.border = g2_border
    g2_cell.alignment = CENTER
    g2_cell.fill = FILL_GREEN
    g2_cell.number_format = "@"

    for start_number, numbering_row in zip((1, 6, 11, 16), (3, 30, 57, 84)):
        for offset, column in enumerate(("G", "H", "I", "J", "K")):
            cell = sheet[f"{column}{numbering_row}"]
            cell.value = start_number + offset
            cell.font = MALGUN_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.fill = FILL_BLUE if column == "K" else FILL_GREEN

    _style_image_slot_block(
        sheet,
        start_row=4,
        end_row=29,
        columns=("G", "H", "I", "J", "K"),
        first_col_merged=False,
    )
    _style_image_slot_block(
        sheet,
        start_row=31,
        end_row=56,
        columns=("G", "H", "I", "J", "K"),
        first_col_merged=True,
    )
    _style_image_slot_block(
        sheet,
        start_row=58,
        end_row=83,
        columns=("G", "H", "I", "J", "K"),
        first_col_merged=True,
    )
    _style_image_slot_block(
        sheet,
        start_row=85,
        end_row=110,
        columns=("G", "H", "I", "J", "K"),
        first_col_merged=True,
    )

    sheet.merge_cells(config.image_header_range)
    for merge_range in config.image_slot_merge_ranges:
        sheet.merge_cells(merge_range)


def _style_image_slot_block(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    columns: tuple[str, ...],
    first_col_merged: bool,
) -> None:
    for col in columns:
        is_first_col_unmerged = col == columns[0] and not first_col_merged
        for row_number in range(start_row, end_row + 1):
            cell = sheet[f"{col}{row_number}"]
            cell.font = CALIBRI_FONT
            cell.fill = FILL_WHITE

            if is_first_col_unmerged:
                top = THIN_SIDE if row_number == start_row else Side(style=None)
                bottom = THIN_SIDE if row_number == end_row else Side(style=None)
                cell.border = Border(
                    left=THIN_SIDE, right=THIN_SIDE, top=top, bottom=bottom
                )
                cell.alignment = VCENTER
            else:
                if row_number == start_row:
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    cell.fill = FILL_WHITE
                    cell.font = MALGUN_FONT


def _create_guide_sheet(workbook: Workbook, config: AppConfig) -> None:
    guide_sheet: Worksheet = workbook.create_sheet(title=config.guide_sheet_name)
    guide_sheet.sheet_view.showGridLines = False
    guide_sheet.sheet_format.defaultColWidth = 11.1667
    guide_sheet.sheet_format.baseColWidth = 8

    for column, width in {
        "A": 8.35156,
        "B": 9.85156,
        "C": 31.1719,
        "D": 39.8516,
        "E": 31.1719,
        "F": 11.1719,
    }.items():
        guide_sheet.column_dimensions[column].width = width

    row_heights = {
        1: 18.0,
        2: 18.0,
        3: 34.5,
        4: 34.5,
        5: 34.5,
        6: 43.5,
        7: 43.5,
        8: 24.75,
        9: 24.0,
        10: 24.0,
        11: 18.0,
        12: 18.0,
        13: 18.0,
        14: 18.0,
        15: 18.0,
        16: 18.0,
        17: 18.0,
        18: 18.0,
        19: 18.0,
        20: 99.0,
    }
    for row_number, height in row_heights.items():
        guide_sheet.row_dimensions[row_number].height = height

    for row_number in range(1, 21):
        for column in "ABCDE":
            cell = guide_sheet[f"{column}{row_number}"]
            cell.font = CALIBRI_FONT
            cell.border = THIN_BORDER
            cell.fill = FILL_WHITE
            cell.alignment = VCENTER

    # row 1
    for column in ("B", "C", "D"):
        cell = guide_sheet[f"{column}1"]
        cell.font = CALIBRI_FONT

    # row 2
    for coord, value in (("B2", "비목"), ("C2", "세목"), ("D2", "비고")):
        cell = guide_sheet[coord]
        cell.value = value
        cell.font = MALGUN_FONT
        cell.border = THIN_BORDER
        cell.fill = FILL_GREEN
        cell.alignment = VCENTER
        cell.number_format = "@"

    _write_guide_data_rows(guide_sheet, config)

    # row 11
    for column in ("B", "C", "D"):
        cell = guide_sheet[f"{column}11"]
        cell.font = CALIBRI_FONT
        cell.border = THIN_BORDER
        cell.fill = FILL_WHITE
        cell.alignment = VCENTER

    _write_guide_notice_section(guide_sheet, config)

    for merged_range in (
        "B3:B5",
        "D3:D5",
        "B6:B7",
        "D6:D7",
        "B9:B10",
        "D9:D10",
        "B12:D12",
        "B13:D20",
    ):
        guide_sheet.merge_cells(merged_range)


def _write_guide_data_rows(sheet: Worksheet, config: AppConfig) -> None:
    guide_data: list[tuple[int, str | None, str | None, str | None, bool]] = [
        (
            3,
            "자기관리",
            "의료/치료",
            "치과치료, 물리치료 등 일반 의료 목적에 사용한 비용 청구 가능\n여러 개월에 걸쳐 사용 가능한 회원권 (예: 헬스장 1년권)의 경우 매 개월마다, 전체 금액을 사용 개월 수로 균등하게 분할한 금액을 회차별로 나누어 신청",
            True,
        ),
        (4, None, "헤어", None, False),
        (5, None, "운동", None, False),
        (
            6,
            "식비",
            "음식점",
            "음식점 사용시 일반음식점만 청구 가능\n주류는 청구 불가. 주류 금액은 제외하여 청구.\n사내 다른 팀원과 함께 이용한 금액에 대해서는 N할로 분할하여 청구하고, 인보이스(영수증/구매내역) 제출 필요",
            True,
        ),
        (7, None, "배달", None, False),
        (8, "통신비", "통신비", "개인 및 업무 장비에 사용한 통신요금 포함", False),
        (
            9,
            "교통비",
            "대중교통",
            "택시의 경우 다른 회사 구성원과 함께에 탑승한 경우, 해당 동승인 인원만큼 나누어 개인별로 청구",
            True,
        ),
        (10, None, "택시", None, False),
    ]

    for row_num, cat, subcat, note, note_wrap in guide_data:
        if cat is not None:
            b_cell = sheet[f"B{row_num}"]
            b_cell.value = cat
            b_cell.font = MALGUN_FONT
            b_cell.border = THIN_BORDER
            b_cell.fill = FILL_WHITE
            b_cell.alignment = CENTER
            b_cell.number_format = "@"

        c_cell = sheet[f"C{row_num}"]
        if subcat is not None:
            c_cell.value = subcat
        c_cell.font = MALGUN_FONT
        c_cell.border = THIN_BORDER
        c_cell.fill = FILL_WHITE
        c_cell.alignment = CENTER
        c_cell.number_format = "@"

        if note is not None:
            d_cell = sheet[f"D{row_num}"]
            d_cell.value = note
            d_cell.font = MALGUN_FONT
            d_cell.border = THIN_BORDER
            d_cell.fill = FILL_WHITE
            d_cell.alignment = LEFT_WRAP if note_wrap else VCENTER
            d_cell.number_format = "@"


def _write_guide_notice_section(sheet: Worksheet, config: AppConfig) -> None:
    # row 12
    b12 = sheet["B12"]
    b12.value = "유의사항"
    b12.font = MALGUN_FONT
    b12.border = THIN_BORDER
    b12.fill = FILL_GREEN
    b12.alignment = CENTER
    b12.number_format = "@"

    # row 13
    b13 = sheet["B13"]
    b13.value = GUIDE_NOTICE_TEXT
    b13.font = MALGUN_FONT
    b13.border = THIN_BORDER
    b13.fill = FILL_WHITE
    b13.alignment = LEFT_WRAP
    b13.number_format = "@"


def _write_rows(sheet: Worksheet, rows: list[ExportRow], config: AppConfig) -> None:
    for row_number in range(config.table_start_row, config.table_border_end_row + 1):
        for column in (
            config.category_column,
            config.subcategory_column,
            config.amount_column,
        ):
            sheet[f"{column}{row_number}"] = None

    max_rows = config.numbered_table_end_row - config.table_start_row + 1
    for index, export_row in enumerate(rows[:max_rows], start=1):
        row_number = config.table_start_row + index - 1
        cat_cell = sheet[f"{config.category_column}{row_number}"]
        cat_cell.value = export_row.category
        cat_cell.alignment = LEFT_VCENTER
        cat_cell.number_format = "@"

        sub_cell = sheet[f"{config.subcategory_column}{row_number}"]
        sub_cell.value = export_row.subcategory
        sub_cell.alignment = LEFT_VCENTER
        sub_cell.number_format = "@"

        sheet[f"{config.amount_column}{row_number}"] = int(export_row.amount)


def _embed_receipt_images(
    sheet: Worksheet, receipts: list[UploadedReceipt], config: AppConfig
) -> None:
    if hasattr(sheet, "_images"):
        cast(Any, sheet)._images = []

    for receipt, anchor in zip(
        receipts[: config.max_receipt_images], config.image_anchor_cells
    ):
        image = open_image_from_bytes(receipt.image_bytes)
        resized = resize_for_excel(
            image,
            max_width=config.image_max_width_px,
            max_height=config.image_max_height_px,
        )
        excel_image = XLImage(BytesIO(image_to_png_bytes(resized)))
        excel_image.anchor = anchor
        sheet.add_image(excel_image)


def _coerce_report_month(report_month_text: date | str) -> date | str:
    if isinstance(report_month_text, datetime):
        return report_month_text.date().replace(day=1)
    if isinstance(report_month_text, date):
        return report_month_text.replace(day=1)

    report_month_value = str(report_month_text).strip()
    for date_format in ("%Y-%m", "%Y-%m-%d"):
        try:
            return (
                datetime.strptime(report_month_value, date_format).date().replace(day=1)
            )
        except ValueError:
            continue
    return report_month_value


def _coerce_rows(rows: list[ExportRow | Mapping[str, object]]) -> list[ExportRow]:
    normalized: list[ExportRow] = []
    for index, row in enumerate(rows, start=1):
        if isinstance(row, ExportRow):
            normalized.append(row)
            continue

        amount_value = row.get("amount")
        if amount_value in (None, ""):
            continue

        category = str(row.get("category") or "기타").strip()
        subcategory = str(row.get("subcategory") or "기타").strip()
        vendor = _optional_str(row.get("vendor"))
        notes = _optional_str(row.get("notes"))
        receipt_date = row.get("receipt_date")

        normalized.append(
            ExportRow(
                number=index,
                category=category,
                subcategory=subcategory,
                amount=Decimal(str(amount_value).replace(",", "").strip()),
                vendor=vendor,
                receipt_date=receipt_date if isinstance(receipt_date, date) else None,
                notes=notes,
            )
        )

    return normalized


def _optional_str(value: object) -> str | None:
    if value in (None, ""):
        return None
    return str(value)
